import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
import xlsxwriter
from shutil import copyfile

file = input('File Path: ')
extension = os.path.splitext(file)[1]
filename = os.path.splitext(file)[0]
pth = os.path.dirname(file)
newfile = os.path.join(pth, filename + '_2' + extension)
df = pd.read_excel(file)
colpick = input('Select Column: ')
cols = list(set(df[colpick].values))

alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

font = Font(name='Calibri',
                 size=12,
                 bold=False,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='FF000000')
fill = PatternFill(fill_type=None,
                 start_color='FFFFFFFF',
                 end_color='FF000000')
border = Border(left=Side(border_style=None,
                           color='FF000000'),
                 right=Side(border_style=None,
                            color='FF000000'),
                 top=Side(border_style=None,
                          color='FF000000'),
                 bottom=Side(border_style=None,
                             color='FF000000'),
                 diagonal=Side(border_style=None,
                               color='FF000000'),
                 diagonal_direction=0,
                 outline=Side(border_style=None,
                              color='FF000000'),
                 vertical=Side(border_style=None,
                               color='FF000000'),
                 horizontal=Side(border_style=None,
                                color='FF000000')
                )
alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=False,
                     shrink_to_fit=False,
                     indent=0)

highlight = NamedStyle(name="highlight")
bd = Side(style='thin', color="000000")

highlight.font = font
highlight.border = Border(left=bd,right=bd,top=bd,bottom=bd)

def sendtofile(cols, deleteCol):
    colpath = pth + "/" + colpick
    if not os.path.exists(colpath):
        os.makedirs(colpath)
    for i in cols:
        tempDf = df[df[colpick] == i]
        if deleteCol == 'y':
            tempDf = tempDf.drop(colpick, axis=1)
        tempDf = tempDf.sort_values(by=['Section', 'ID'])
        tempFilePath ="{}/{}.xlsx".format(colpath, i) 
        tempDf.to_excel(tempFilePath, sheet_name=i, index=False)
        wb = load_workbook(tempFilePath)
        wb.add_named_style(highlight)
        ws = wb[i]
        rows, columns = tempDf.shape
        for c in range(columns):
            for r in range(1,rows):
                ws[f'{alphabet[c]}{r}'].style=highlight
            
        wb.save(tempFilePath)
    
    print('\nCompleted')
    print('Thanks for using this program.')
    return


def sendtosheet(cols):
    copyfile(file, newfile)
    for j in cols:
        writer = pd.ExcelWriter(newfile, engine='openpyxl')
        for myname in cols:
            mydf = df.loc[df[colpick] == myname]
            mydf.to_excel(writer, sheet_name=myname, index=False)
        writer.save()

    print('\nCompleted')
    print('Thanks for using this program.')
    return


print(
    'You data will split based on these values {} and create {} files or sheets based on next selection. If you are ready to proceed please type "Y" and hit enter. Hit "N" to exit.'
    .format(', '.join(cols), len(cols)))
while True:
    x = input('Ready to Proceed (Y/N): ').lower()
    if x == 'y':
        while True:
            s = input('Split into different Sheets or File (S/F): ').lower()
            deleteCol = input(f'Delete  the column {colpick} (Y/N): ').lower()
            if s == 'f':
                sendtofile(cols,deleteCol)
                break
            elif s == 's':
                sendtosheet(cols)
                break
            else:
                continue
        break
    elif x == 'n':
        print('\nThanks for using this program.')
        break

    else:
        continue
